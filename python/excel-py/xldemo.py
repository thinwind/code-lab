# -*- coding: utf-8 -*-
"""
@Time   :  2023-Sep-08 11:38
@Author :  Shang Yehua
@Email  :  niceshang@outlook.com
@Desc   :  
        解析excel文件 DEMO
"""
from openpyxl import load_workbook
from openpyxl import Workbook

# 读取excel文件
wb = load_workbook('demo.xlsx')
# 获取所有sheet名称
print(wb.sheetnames)
# 获取指定sheet
sheet = wb['样本案例']

# 获取指定单元格的值
print(sheet['B2'].value)
print(sheet['B1'].value)
print(sheet['B3'].value)
print(sheet['B4'].value)

# 遍历workbook 中的 sheet
for sheet in wb:
    print(sheet.title)
    
# copy sheet
source = wb['问题清单']
target = wb.copy_worksheet(source)

